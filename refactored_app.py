import openpyxl

def load_excel_data(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    return wb, sheet

def find_start_and_end_rows(sheet):
    first, start_row, end_row = [cell.row for cell in sheet['A'] if cell.value in ["Total", "REGIÓN"]]
    end_row = start_row + 2
    return start_row, end_row

def identify_weekday_and_weekend_columns(sheet):
    weekday_col_start, weekend_col_start = [cell.column for cell in sheet['11'] if cell.value in ["Lunes a Viernes", "Sábado-Domingo"]]
    return weekday_col_start, weekend_col_start

def extract_station_names(sheet, start_row, end_row, weekday_col_start):
    station_name = [sheet.cell(row, weekday_col_start - 1).value for row in range(start_row, end_row)]
    return station_name

def extract_time_stamps(sheet, start_row, weekday_col_start):
    time_stamp = [sheet.cell(start_row - 1, col).value for col in range(weekday_col_start + 2, sheet.max_column + 1) 
                  if sheet.cell(start_row - 1, col).value not in ['Total', 'TOTAL']]
    return time_stamp

def process_data(sheet, start_row, end_row, weekday_col_start, weekend_col_start, station_name, time_stamp):
    nested_aud = []
    for data in [[weekday_col_start, weekend_col_start],[weekend_col_start, sheet.max_column + 1]]:
        if data[0] == weekday_col_start:
            flag = 0
        elif data[0] == weekend_col_start:
            flag = 1
        for row in range(start_row, end_row):
            audience = []
            mapped_dict = {}
            for col in range(data[0] + 2, data[1]):           
                audience.append(sheet.cell(row, col).value)
            for aud, time_data in zip(audience, time_stamp):
                new_time = time_data.split(".")
                new_time = int(new_time[0])
                if new_time == 24:
                    new_time = 0
                new_time = f'{new_time}.00'
                if new_time in mapped_dict:
                    mapped_dict[new_time] += aud
                else:
                    mapped_dict[new_time] = {}
                    mapped_dict[new_time] = aud
            nested_aud.append({"Tv_Program_Channel": station_name[row - start_row], "Time": mapped_dict, "Flag": flag})
    return nested_aud

def generate_sql_insert_statements(nested_aud, day_mapping):
    
    sql_statements = []
    for i in nested_aud:
        if i["Flag"] == 0:
            for range_data in range(0, 5):
                day = day_mapping.get(range_data, "Unknown")
                sql_statement = f'INSERT INTO radio_audience(station_name, timestamp, audience, flag, day_of_week) VALUES ({i["Tv_Program_Channel"]},{i["Time"]},{i["Flag"]},{day});'
                sql_statements.append(sql_statement)
        else:
            for range_data in range(5, 7):
                day = day_mapping.get(range_data, "Unknown")
                sql_statement = f'INSERT INTO radio_audience(station_name, timestamp, audience, flag, day_of_week) VALUES ({i["Tv_Program_Channel"]},{i["Time"]},{i["Flag"]},{day});'
                sql_statements.append(sql_statement)
    return sql_statements

if __name__ == "__main__":
    wb, sheet = load_excel_data('Excel_Data_Reader/spain_data.xlsx', 'Hoja2')
    start_row, end_row = find_start_and_end_rows(sheet)
    weekday_col_start, weekend_col_start = identify_weekday_and_weekend_columns(sheet)
    station_name = extract_station_names(sheet, start_row, end_row, weekday_col_start)
    time_stamp = extract_time_stamps(sheet, start_row, weekday_col_start)
    nested_aud = process_data(sheet, start_row, end_row, weekday_col_start, weekend_col_start, station_name, time_stamp)
    day_mapping = {0: "Monday", 1: "Tuesday", 2: "Wednesday", 3: "Thursday", 4: "Friday", 5: "Saturday", 6: "Sunday"}
    sql_statements = generate_sql_insert_statements(nested_aud, day_mapping)
    for statement in sql_statements:
        print(statement)
