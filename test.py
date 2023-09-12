import openpyxl

wb = openpyxl.load_workbook('spain_data.xlsx')
sheet = wb['Hoja2']

total, start_row, end_row = [cell.row for cell in sheet['A'] if cell.value in ["Total", "REGIÓN"]]
end_row =start_row+2

weekday_col_start, weekend_col_start = [cell.column for cell in sheet['11'] if cell.value in ["Lunes a Viernes", "Sábado-Domingo"]]

station_name = [sheet.cell(row, weekday_col_start - 1).value for row in range(start_row, end_row)]

time_stamp = [sheet.cell(start_row - 1, col).value for col in range(weekday_col_start + 2, sheet.max_column + 1) 
              if sheet.cell(start_row - 1, col).value not in ['Total', 'TOTAL']]

nested_aud = []
for data in [[weekday_col_start, weekend_col_start],[weekend_col_start, sheet.max_column+1]]:
    if data[0] == weekday_col_start:
        flag = 0
    elif data[0] == weekend_col_start:
        flag = 1
    for row in range(start_row,end_row):
        audience = []
        mapped_dict = {}
        for col in range (data[0]+2,data[1]):           
            audience.append(sheet.cell(row,col).value)
        for aud,time_data in zip(audience,time_stamp):
            new_time = time_data.split(".")
            new_time = int(new_time[0])
            new_time = f'{new_time}.00-{new_time+1}.00'
            if new_time in mapped_dict:
                mapped_dict[new_time] += aud
            else:
                mapped_dict[new_time] = {}
                mapped_dict[new_time] = aud
        nested_aud.append({"Tv_Program_Channel": station_name[row - start_row], "Time": mapped_dict, "Flag":flag})

day_mapping = {0: "Monday",1: "Tuesday",2: "Wednesday",3: "Thursday",4: "Friday",5: "Saturday",6: "Sunday"}

for i in nested_aud:
    if i["Flag"] == 0:
        for range_data in range(0,5):
            day = day_mapping.get(range_data, "Unknown")
            op=f'Insert into radio_audience(station_name, timestamp, audience, flag, day_of_week) values ({i["Tv_Program_Channel"]},{i["Time"]},{i["Flag"]},{day});'
            print(op)
    else:
        for range_data in range(5,7):
            day = day_mapping.get(range_data, "Unknown")
            op=f'Insert into radio_audience(station_name, timestamp, audience, flag, day_of_week) values ({i["Tv_Program_Channel"]},{i["Time"]},{i["Flag"]},{day});'
            print(op)
 